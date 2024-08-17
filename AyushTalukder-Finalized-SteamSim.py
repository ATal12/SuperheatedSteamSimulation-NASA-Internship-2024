import openpyxl
from openpyxl import Workbook
#from openpyxl import load_workbook
import os
import csv
import math
from datetime import datetime
# load the excel file
#rb = load_workbook(filename="C:\Users\Ayush\Downloads\Steam_Tables_1.23.xlsx")

# open the superheated steam sheet
#w_sheet = rb.active
#sheet=rb["compressed_liq_and_superh_steam"]

def main():
  print("ran main")
  airship_vol=280000
  steam_vol=airship_vol/3
  steam_sim(530,5,285,23,350,airship_vol, steam_vol, 6051.8, 235000,0.5,2276.65)
  

def getMass(temp, pressure,vol):
  if(pressure<=0.01):
    pressure=0.01 #minimum possible pressure value, needs rounding
  #print("ran method")
  #print("pressure: "+str(pressure))
 # print("temperature: "+str(temp))
  vol_liters=vol*1000 #convert from m^3 to liters
  filename=r"C:\Users\Ayush\Downloads\compressed_liquid_and_superheated_steam_V1.3.csv"
  tempVal=1.0
  pressureVal=1.0
  
  percentDiff=0.05 #find a value within percentDiff of the true value; 0.01 would mean 1%, 0.02 would mean 2%, etc...
  with open(filename, mode='r') as file:
    
    csvFile = csv.reader(file)
    for x in range(7):
      next(csvFile)
    for lines in file:
      values = lines.strip().split(',')
      pressureVal=float(values[0])
      tempVal=float(values[1])
      if abs(pressureVal-pressure)/abs((pressureVal+pressure/2))<=percentDiff and abs(tempVal-temp)/abs((tempVal+temp)/2)<=percentDiff:
       # print(str(pressure)+" is the inputted pressure, and "+str(pressureVal)+" is the rounded one")
        #print(str(temp)+" is the inputted temp, and "+str(tempVal)+" is the rounded one")
        #print(float(values[3]))
        #now calculate moles using pv=nrt, calculate molecular mass using 3R/C*1000 and calculate mass=mol*molmass
       # print("specific heat: "+values[6])
        #massVal=pressure*vol_liters/(8.314*(temp+273.15))*3*8.314/float(values[6])*1000
      
        #print("mass with thermodynamics equation: "+str(massVal))
        massVal=vol*float(values[3])
       # print("mass with density*volume: " +str(massVal))
        return float(massVal)        
#dur is time duration of simulation in seconds; t_int is the time interval used for the simulation; stm_temp is the initial temperature of the steam;
#surr_temp is initial temperature of surroundings; alt is initial altitude above sea level; v_airship is the total volume of the airship; v_steam is the volume of steam used;
#s_area_ball is the surface area of the ballonets; m_car is the mass of cargo; pressure_ball is the pressure inside the ballonets

  
def steam_sim(dur,t_int,stm_temp,surr_temp,alt,v_airship, v_steam,s_area_ball,m_car,pressure_ball,cross_sec_area):
  #establishing necessary constant and initial values
  print("started simulation")
  velocity_final=0
  alt_init=alt
  surround_temp_init=surr_temp
  density_helium=0.166 #we are using a constant density of helium
  vol_helium=v_airship-v_steam #assume airship  filled fully with steam/helium
  mass_helium=density_helium*vol_helium #mass=d*v

  #print("mass of helium:" +str(mass_helium))
  wb=Workbook()
  #ws=wb.active()
  steamsheet=wb.create_sheet("Steam Simulation")
  steamsheet.cell(1,1,"time (s)")
  steamsheet.cell(1,2,"temperature (C)")
  steamsheet.cell(1,3,"altitude (m)")
  steamsheet.cell(1,4,"difference between helium and steam forces (N)")
  steamsheet.cell(1,5,"steam force percentage increase relative to helium force")
  row=2
  for time in range(0, dur+1, t_int):
    #set initial velocity
    velocity_init=velocity_final

    #update pressure, surrounding air density and steam density
    pressure_kpa_outside=101325*(1-2.25577*10**-5*alt)**5.25588/1000
    density_surround=pressure_kpa_outside*1000*4.81*10**-26/(1.380649*10**-23*(surr_temp+273.15))
    #sheet["C8"] = pressure_kpa/1000
    #sheet["D8"]=stm_temp
    
    #print(sheet["E8"].value)
    #Float = float(sheet["E8"])
    #print("1/density_steam = "+str(sheet["E8"]))
    mass_steam= getMass(stm_temp,pressure_ball/1000,v_steam)
    #print("steam mass: " + str(mass_steam))

    #calculate all steam motion values: net force, steam temperature loss, altitude change, velocity change 

    #net force - NOTE: NEED TO FIND A WAY TO INCORPORATE DRAG FORCE
    force_buoy=density_surround*v_airship*9.8
    weight_total=mass_steam*9.8+m_car*9.8+mass_helium*9.8
    print(velocity_init)
    force_drag=1/2*0.045*velocity_init**2*density_surround*cross_sec_area
    print(force_drag)
    force_net=force_buoy-weight_total-force_drag
    #print(str(force_buoy)+" is the buoyant force and "+str(weight_total)+" is the weight carried.")
    #the rest
    steam_temp_change=0.13*s_area_ball*(stm_temp-surr_temp)/(mass_steam*2100)
    print(str(steam_temp_change)+" is the change in steam temperature.")
    alt_change=velocity_init*t_int+1/2*force_net/(mass_steam+mass_helium+m_car)*t_int**2
    velocity_final=velocity_init+force_net/(mass_steam+mass_helium+m_car)*t_int
    surr_temp= surround_temp_init-9.8/1000*(alt-alt_init)

    #calculating difference between net force when using steam and when using helium
    force_net_helium=9.8*v_airship*density_surround-(density_helium*v_steam+m_car+mass_helium)*9.8-1/2*0.045*velocity_init**2*density_surround*cross_sec_area
    diff_between_steam_helium=force_net-force_net_helium
    percentDiff=diff_between_steam_helium/abs((force_net_helium+force_net)/2)*100

    #updating temperature and altitude values
    alt+=alt_change
    print("altitude change: "+str(alt_change))
    stm_temp-=steam_temp_change
    print("time elapsed: "+str(time))
    print("altitude: "+str(alt))
    print("steam temperature: "+str(stm_temp))
    print("steam is currently producing "+str(diff_between_steam_helium)+"N more force than helium would with the same volume used.")
    steamsheet.cell(row,1,time)
    steamsheet.cell(row,2,stm_temp)
    steamsheet.cell(row,3,alt)
    steamsheet.cell(row,4,diff_between_steam_helium)
    steamsheet.cell(row,5,percentDiff)
    row+=1
    
  #save the file
  time=str(datetime.now())[4:10]+"-"+str(datetime.now())[11:13]+str(datetime.now())[14:16]
  filename="C:\\Users\\Ayush\\Downloads\\SteamSimulationExcelSheets\\SteamSim_"+time+".xlsx"
  print(filename)
  wb.save(filename)


main() #calls main
